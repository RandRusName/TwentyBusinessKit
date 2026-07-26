from __future__ import annotations

import base64
import hashlib
import io
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .generator import DocumentGenerationError, GeneratedLocalFile, XLSX_CONTENT_TYPE


A1_CELL_REGEX = re.compile(r"^\$?([A-Za-z]+)\$?([1-9][0-9]*)$")
MAX_TEMPLATE_BYTES = 8 * 1024 * 1024
XLSX_MAGIC = b"PK"
PREVIEW_MAX_ROWS = 80
PREVIEW_MAX_COLUMNS = 30
PREVIEW_MAX_DISPLAY_LENGTH = 120


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _decode_base64_file(template_file_base64: str) -> bytes:
    if not isinstance(template_file_base64, str) or template_file_base64.strip() == "":
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateFileBase64 is required")
    try:
        data = base64.b64decode(template_file_base64, validate=True)
    except Exception as error:
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateFileBase64 must be valid base64") from error
    if len(data) == 0:
        raise DocumentGenerationError("PAYLOAD_INVALID", "template file is empty")
    if len(data) > MAX_TEMPLATE_BYTES:
        raise DocumentGenerationError("PAYLOAD_TOO_LARGE", "template file exceeds size limit")
    if not data.startswith(XLSX_MAGIC):
        raise DocumentGenerationError("PAYLOAD_INVALID", "template file must be an XLSX (ZIP) package")
    return data


def _assert_xlsx_filename(original_file_name: str) -> None:
    if not isinstance(original_file_name, str) or original_file_name.strip() == "":
        raise DocumentGenerationError("PAYLOAD_INVALID", "originalFileName is required")
    lowered = original_file_name.strip().lower()
    if not lowered.endswith(".xlsx"):
        raise DocumentGenerationError("PAYLOAD_INVALID", "originalFileName must end with .xlsx")
    if lowered.endswith(".xlsm"):
        raise DocumentGenerationError("PAYLOAD_INVALID", "macro-enabled templates are not supported for custom XLSX")


def _open_workbook(data: bytes) -> Workbook:
    try:
        workbook = load_workbook(filename=io.BytesIO(data), data_only=False, keep_vba=False)
    except Exception as error:
        message = str(error).lower()
        if "password" in message or "encrypted" in message:
            raise DocumentGenerationError("PAYLOAD_INVALID", "password-protected workbooks are not supported") from error
        raise DocumentGenerationError("PAYLOAD_INVALID", "unable to open XLSX workbook") from error
    return workbook


def _visible_sheets(workbook: Workbook) -> list[Worksheet]:
    sheets: list[Worksheet] = []
    for sheet in workbook.worksheets:
        state = getattr(sheet, "sheet_state", "visible")
        if state == "hidden" or state == "veryHidden":
            continue
        sheets.append(sheet)
    return sheets


def _truncate_display(value: str) -> str:
    if len(value) <= PREVIEW_MAX_DISPLAY_LENGTH:
        return value
    return f"{value[: PREVIEW_MAX_DISPLAY_LENGTH - 1]}…"


def _preview_cell_payload(
    *,
    row: int,
    column: int,
    raw_value: Any,
    is_merged: bool,
) -> dict[str, Any]:
    address = f"{get_column_letter(column)}{row}"
    has_formula = isinstance(raw_value, str) and raw_value.startswith("=")
    value: Any
    if raw_value is None:
        value = None
        display = ""
    elif isinstance(raw_value, bool):
        value = raw_value
        display = "TRUE" if raw_value else "FALSE"
    elif isinstance(raw_value, (int, float)):
        value = raw_value
        display = str(raw_value)
    else:
        text = str(raw_value)
        value = text if not has_formula else text
        display = text
    return {
        "row": row,
        "column": column,
        "address": address,
        "value": value,
        "displayValue": _truncate_display(display),
        "hasFormula": has_formula,
        "isMerged": is_merged,
    }


def _build_sheet_preview(sheet: Worksheet) -> dict[str, Any]:
    max_row = min(int(sheet.max_row or 1), PREVIEW_MAX_ROWS)
    max_column = min(int(sheet.max_column or 1), PREVIEW_MAX_COLUMNS)
    merged_coords: set[tuple[int, int]] = set()
    for merged in sheet.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                if row <= max_row and column <= max_column:
                    merged_coords.add((row, column))

    cells: list[dict[str, Any]] = []
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            cell = sheet.cell(row=row, column=column)
            raw_value = cell.value
            is_merged = (row, column) in merged_coords
            if raw_value is None and not is_merged:
                continue
            cells.append(
                _preview_cell_payload(
                    row=row,
                    column=column,
                    raw_value=raw_value,
                    is_merged=is_merged,
                )
            )

    return {
        "rowLimit": max_row,
        "columnLimit": max_column,
        "cells": cells,
    }


def inspect_xlsx(*, template_file_base64: str, original_file_name: str) -> dict[str, Any]:
    _assert_xlsx_filename(original_file_name)
    data = _decode_base64_file(template_file_base64)
    workbook = _open_workbook(data)
    visible = _visible_sheets(workbook)
    if len(visible) == 0:
        raise DocumentGenerationError("PAYLOAD_INVALID", "workbook must contain at least one visible sheet")

    named_by_sheet: dict[str, list[dict[str, str]]] = {sheet.title: [] for sheet in workbook.worksheets}
    defined_names = getattr(workbook, "defined_names", None)
    if defined_names is not None:
        values = []
        if hasattr(defined_names, "values"):
            values = list(defined_names.values())
        elif hasattr(defined_names, "definedName"):
            values = list(defined_names.definedName)
        for defined_name in values:
            name = getattr(defined_name, "name", None)
            refers_to = getattr(defined_name, "attr_text", None) or getattr(defined_name, "value", "") or ""
            if not isinstance(name, str) or name.strip() == "":
                continue
            entry = {"name": name, "refersTo": str(refers_to)}
            destinations = []
            try:
                destinations = list(defined_name.destinations)
            except Exception:
                destinations = []
            if destinations:
                for sheet_title, _ in destinations:
                    named_by_sheet.setdefault(sheet_title, []).append(entry)
            else:
                for sheet_title in named_by_sheet:
                    named_by_sheet[sheet_title].append(entry)

    sheets_meta = []
    for sheet in visible:
        merges = [str(range_) for range_ in sheet.merged_cells.ranges]
        max_row = sheet.max_row or 1
        max_column = sheet.max_column or 1
        sheets_meta.append(
            {
                "name": sheet.title,
                "maxRow": int(max_row),
                "maxColumn": int(max_column),
                "mergedRanges": merges,
                "namedRanges": named_by_sheet.get(sheet.title, []),
                "preview": _build_sheet_preview(sheet),
            }
        )

    return {
        "status": "success",
        "workbook": {"sheets": sheets_meta},
        "sha256": _sha256(data),
    }


def _resolve_path(payload: dict[str, Any], field_path: str) -> Any:
    current: Any = payload
    for part in field_path.split("."):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def _normalize_a1(cell: str) -> tuple[str, int]:
    match = A1_CELL_REGEX.fullmatch(cell.strip())
    if match is None:
        raise DocumentGenerationError("PAYLOAD_INVALID", f"invalid A1 cell '{cell}'")
    column = match.group(1).upper()
    row = int(match.group(2))
    return column, row


def _write_cell_value(sheet: Worksheet, cell_ref: str, value: Any) -> None:
    cell = sheet[cell_ref]
    if value is None:
        cell.value = None
        return
    if isinstance(value, bool):
        cell.value = value
        return
    if isinstance(value, (int, float)):
        cell.value = value
        return
    cell.value = str(value)


def _copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def _row_has_vertical_merge(sheet: Worksheet, row_index: int) -> bool:
    for merged in sheet.merged_cells.ranges:
        min_row, max_row = merged.min_row, merged.max_row
        if min_row != max_row and min_row <= row_index <= max_row:
            return True
    return False


def _sheet_has_table_overlapping_row(sheet: Worksheet, row_index: int) -> bool:
    tables = getattr(sheet, "tables", {}) or {}
    for table in tables.values():
        ref = getattr(table, "ref", None)
        if not ref:
            continue
        # e.g. "A10:G20"
        try:
            start, end = str(ref).split(":")
            _, start_row = coordinate_from_string(start)
            _, end_row = coordinate_from_string(end)
        except Exception:
            continue
        if start_row <= row_index <= end_row:
            return True
    return False


def _assert_safe_template_row(sheet: Worksheet, template_row: int) -> None:
    if _row_has_vertical_merge(sheet, template_row):
        raise DocumentGenerationError(
            "PAYLOAD_INVALID",
            f"template row {template_row} intersects a vertical merged range; unsupported in MVP",
        )
    if _sheet_has_table_overlapping_row(sheet, template_row):
        raise DocumentGenerationError(
            "PAYLOAD_INVALID",
            f"template row {template_row} intersects an Excel Table; unsupported in MVP",
        )


def _copy_horizontal_merges(sheet: Worksheet, source_row: int, target_row: int) -> None:
    to_add = []
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row == merged.max_row == source_row:
            start = f"{get_column_letter(merged.min_col)}{target_row}"
            end = f"{get_column_letter(merged.max_col)}{target_row}"
            to_add.append(f"{start}:{end}")
    for ref in to_add:
        sheet.merge_cells(ref)


def _expand_table_binding(
    workbook: Workbook,
    payload: dict[str, Any],
    binding: dict[str, Any],
) -> None:
    sheet_name = binding["sheetName"]
    if sheet_name not in workbook.sheetnames:
        raise DocumentGenerationError("PAYLOAD_INVALID", f"sheet '{sheet_name}' not found")
    sheet = workbook[sheet_name]
    template_row = int(binding["templateRow"])
    _assert_safe_template_row(sheet, template_row)

    collection_path = binding["collectionPath"]
    items = _resolve_path(payload, collection_path)
    if not isinstance(items, list):
        items = []

    min_rows = int(binding.get("minRows", 1))
    render_count = max(len(items), min_rows)
    if render_count == 0:
        return

    copy_style = bool(binding.get("copyStyleFromTemplateRow", True))
    preserve_formulas = bool(binding.get("preserveFormulas", True))
    columns = binding["columns"]

    if render_count > 1:
        sheet.insert_rows(template_row + 1, amount=render_count - 1)

    for offset in range(render_count):
        target_row = template_row + offset
        item = items[offset] if offset < len(items) else {}

        if offset > 0:
            if copy_style:
                source_height = sheet.row_dimensions[template_row].height
                if source_height is not None:
                    sheet.row_dimensions[target_row].height = source_height
                max_col = max(sheet.max_column or 1, 1)
                for col_idx in range(1, max_col + 1):
                    source_cell = sheet.cell(row=template_row, column=col_idx)
                    target_cell = sheet.cell(row=target_row, column=col_idx)
                    _copy_cell_style(source_cell, target_cell)
                    if preserve_formulas and isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                        target_cell.value = _shift_formula_row(source_cell.value, template_row, target_row)
                    elif source_cell.value is not None:
                        # Copy static template-row values into inserted rows, then overwrite bindings.
                        target_cell.value = source_cell.value
            _copy_horizontal_merges(sheet, template_row, target_row)

        for column in columns:
            field_path = column["fieldPath"]
            column_letter, column_row = _normalize_a1(column["cell"])
            if column_row != template_row:
                raise DocumentGenerationError(
                    "PAYLOAD_INVALID",
                    f"table column cell row must equal templateRow ({template_row})",
                )
            cell_ref = f"{column_letter}{target_row}"
            value = item.get(field_path) if isinstance(item, dict) else None
            _write_cell_value(sheet, cell_ref, value)


def _shift_formula_row(formula: str, source_row: int, target_row: int) -> str:
    if source_row == target_row:
        return formula
    try:
        from openpyxl.formula.translate import Translator

        origin = f"A{source_row}"
        destination = f"A{target_row}"
        return Translator(formula, origin=origin).translate_formula(destination)
    except Exception:
        return formula


def _apply_scalar_bindings(
    workbook: Workbook,
    payload: dict[str, Any],
    bindings: list[dict[str, Any]],
) -> None:
    for binding in bindings:
        sheet_name = binding["sheetName"]
        if sheet_name not in workbook.sheetnames:
            raise DocumentGenerationError("PAYLOAD_INVALID", f"sheet '{sheet_name}' not found")
        sheet = workbook[sheet_name]
        column, row = _normalize_a1(binding["cell"])
        cell_ref = f"{column}{row}"
        value = _resolve_path(payload, binding["fieldPath"])
        _write_cell_value(sheet, cell_ref, value)


def _validate_mapping_shape(mapping: dict[str, Any]) -> None:
    if not isinstance(mapping, dict) or mapping.get("schemaVersion") != "1.0":
        raise DocumentGenerationError("PAYLOAD_INVALID", "mapping.schemaVersion must be '1.0'")
    if not isinstance(mapping.get("scalarBindings"), list) or not isinstance(mapping.get("tableBindings"), list):
        raise DocumentGenerationError("PAYLOAD_INVALID", "mapping bindings must be arrays")
    if len(mapping["scalarBindings"]) == 0 and len(mapping["tableBindings"]) == 0:
        raise DocumentGenerationError("PAYLOAD_INVALID", "mapping must contain at least one binding")


def _set_full_calc_on_load(workbook: Workbook) -> None:
    try:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
    except Exception:
        pass


def render_custom_xlsx(
    *,
    template_bytes: bytes,
    payload: dict[str, Any],
    mapping: dict[str, Any],
    output_dir: Path,
    output_file_name: str,
) -> GeneratedLocalFile:
    if not template_bytes.startswith(XLSX_MAGIC):
        raise DocumentGenerationError("PAYLOAD_INVALID", "template file must be an XLSX (ZIP) package")
    if len(template_bytes) > MAX_TEMPLATE_BYTES:
        raise DocumentGenerationError("PAYLOAD_TOO_LARGE", "template file exceeds size limit")

    _validate_mapping_shape(mapping)
    workbook = _open_workbook(template_bytes)
    if len(_visible_sheets(workbook)) == 0:
        raise DocumentGenerationError("PAYLOAD_INVALID", "workbook must contain at least one visible sheet")

    # Expand tables first so scalar writes that reference cells below tables stay correct
    # only when scalars are outside the shifted region; MVP writes scalars after expansion.
    for binding in mapping["tableBindings"]:
        if not isinstance(binding, dict):
            raise DocumentGenerationError("PAYLOAD_INVALID", "table binding must be an object")
        _expand_table_binding(workbook, payload, binding)

    for binding in mapping["scalarBindings"]:
        if not isinstance(binding, dict):
            raise DocumentGenerationError("PAYLOAD_INVALID", "scalar binding must be an object")
        _apply_scalar_bindings(workbook, payload, [binding])

    _set_full_calc_on_load(workbook)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_file_name
    workbook.save(output_path)
    data = output_path.read_bytes()
    return GeneratedLocalFile(
        format="xlsx",
        file_name=output_file_name,
        content_type=XLSX_CONTENT_TYPE,
        size=len(data),
        sha256=_sha256(data),
        path=output_path,
    )


def load_template_bytes_from_render_config(
    render_config: dict[str, Any],
    *,
    storage: Any | None = None,
) -> bytes:
    if not isinstance(render_config, dict) or render_config.get("templateSource") != "custom-xlsx":
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateRenderConfig.templateSource must be custom-xlsx")

    template_file = render_config.get("templateFile")
    if not isinstance(template_file, dict):
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateRenderConfig.templateFile is required")

    content_base64 = template_file.get("contentBase64")
    if isinstance(content_base64, str) and content_base64.strip() != "":
        data = _decode_base64_file(content_base64)
    else:
        storage_key = template_file.get("storageKey")
        if not isinstance(storage_key, str) or storage_key.strip() == "":
            raise DocumentGenerationError("PAYLOAD_INVALID", "templateFile.storageKey is required")
        if storage is None:
            raise DocumentGenerationError("DOCUMENT_STORAGE_FAILED", "storage is required to load custom template")
        data = storage.get_bytes(storage_key=storage_key)

    expected_sha = template_file.get("sha256")
    if not isinstance(expected_sha, str) or not re.fullmatch(r"[0-9a-f]{64}", expected_sha):
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateFile.sha256 must be a sha256 hex digest")
    actual_sha = _sha256(data)
    if actual_sha != expected_sha:
        raise DocumentGenerationError("PAYLOAD_INVALID", "templateFile.sha256 does not match template bytes")
    return data
