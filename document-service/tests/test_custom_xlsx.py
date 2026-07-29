from __future__ import annotations

import base64
import hashlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from mikoton_document_service.custom_xlsx import (
    inspect_xlsx,
    render_custom_xlsx,
    store_xlsx_template,
)
from mikoton_document_service.generator import DocumentGenerationError, LocalDocumentStorage


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "КП"
    sheet["B2"] = "number"
    sheet["B3"] = "company"
    sheet["A15"] = 1
    sheet["B15"] = "item"
    sheet["C15"] = 1
    sheet["D15"] = 100
    sheet["E15"] = "=C15*D15"
    sheet["E15"].font = Font(bold=True)
    sheet["A20"] = "below"
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "template.xlsx"
        workbook.save(path)
        return path.read_bytes()


def _payload() -> dict:
    return {
        "schemaVersion": "2.0",
        "templateCode": "mikoton-commercial-proposal",
        "templateVersion": "2",
        "proposal": {
            "id": "11111111-1111-4111-8111-111111111111",
            "number": "КП-001 от 26.07.2026",
            "title": "Demo",
            "date": "2026-07-26",
            "language": "ru-RU",
            "currencyCode": "RUB",
            "validityDays": 14,
            "amount": 300,
        },
        "customer": {
            "companyId": "33333333-3333-4333-8333-333333333333",
            "companyName": "ООО Заказчик",
            "contactName": "Иван",
        },
        "contractor": {"name": "Mikoton", "email": "a@b.c"},
        "content": {
            "contextAndGoal": "",
            "workItems": [
                {
                    "position": 1,
                    "block": "A",
                    "name": "Item 1",
                    "description": "",
                    "quantity": 1,
                    "unit": "шт",
                    "unitPrice": 100,
                    "discountPercent": 0,
                    "lineAmount": 100,
                    "currencyCode": "RUB",
                },
                {
                    "position": 2,
                    "block": "A",
                    "name": "Item 2",
                    "description": "",
                    "quantity": 2,
                    "unit": "шт",
                    "unitPrice": 100,
                    "discountPercent": 0,
                    "lineAmount": 200,
                    "currencyCode": "RUB",
                },
            ],
            "plan": [],
            "paymentTerms": "",
            "assumptions": "",
            "nextStep": "",
        },
    }


def _mapping() -> dict:
    return {
        "schemaVersion": "1.0",
        "scalarBindings": [
            {
                "kind": "scalar",
                "fieldPath": "proposal.number",
                "sheetName": "КП",
                "cell": "B2",
            },
            {
                "kind": "scalar",
                "fieldPath": "customer.companyName",
                "sheetName": "КП",
                "cell": "B3",
            },
        ],
        "tableBindings": [
            {
                "kind": "table",
                "collectionPath": "content.workItems",
                "sheetName": "КП",
                "templateRow": 15,
                "insertMode": "insertRowsAndShiftDown",
                "minRows": 1,
                "copyStyleFromTemplateRow": True,
                "preserveFormulas": True,
                "columns": [
                    {"fieldPath": "name", "cell": "B15"},
                    {"fieldPath": "quantity", "cell": "C15"},
                    {"fieldPath": "unitPrice", "cell": "D15"},
                ],
            }
        ],
    }


class CustomXlsxTests(unittest.TestCase):
    def test_inspect_xlsx(self) -> None:
        data = _workbook_bytes()
        result = inspect_xlsx(
            template_file_base64=base64.b64encode(data).decode("ascii"),
            original_file_name="demo.xlsx",
        )
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["sha256"], hashlib.sha256(data).hexdigest())
        self.assertEqual(result["workbook"]["sheets"][0]["name"], "КП")

    def test_inspect_includes_preview_cells(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "КП"
        sheet["B2"] = "hello"
        sheet["C3"] = "=B2"
        sheet.merge_cells("A5:B5")
        sheet["A5"] = "merged"
        # Fill enough content to exercise limits without making a huge file.
        for row in range(1, 90):
            sheet.cell(row=row, column=31, value=f"col31-{row}")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "preview.xlsx"
            workbook.save(path)
            data = path.read_bytes()

        result = inspect_xlsx(
            template_file_base64=base64.b64encode(data).decode("ascii"),
            original_file_name="preview.xlsx",
        )
        sheet_meta = result["workbook"]["sheets"][0]
        preview = sheet_meta["preview"]
        self.assertLessEqual(preview["rowLimit"], 80)
        self.assertLessEqual(preview["columnLimit"], 30)
        addresses = {cell["address"]: cell for cell in preview["cells"]}
        self.assertEqual(addresses["B2"]["displayValue"], "hello")
        self.assertTrue(addresses["C3"]["hasFormula"])
        self.assertTrue(addresses["A5"]["isMerged"])
        self.assertTrue(all(cell["column"] <= 30 for cell in preview["cells"]))
        self.assertTrue(all(cell["row"] <= 80 for cell in preview["cells"]))

    def test_scalar_and_table_expansion(self) -> None:
        data = _workbook_bytes()
        with tempfile.TemporaryDirectory() as tmp:
            generated = render_custom_xlsx(
                template_bytes=data,
                payload=_payload(),
                mapping=_mapping(),
                output_dir=Path(tmp),
                output_file_name="out.xlsx",
            )
            workbook = load_workbook(generated.path)
            sheet = workbook["КП"]
            self.assertEqual(sheet["B2"].value, "КП-001 от 26.07.2026")
            self.assertEqual(sheet["B3"].value, "ООО Заказчик")
            self.assertEqual(sheet["B15"].value, "Item 1")
            self.assertEqual(sheet["B16"].value, "Item 2")
            self.assertEqual(sheet["C16"].value, 2)
            self.assertEqual(sheet["D16"].value, 100)
            self.assertTrue(sheet["E16"].value.startswith("="))
            self.assertTrue(sheet["E16"].font.bold)
            self.assertEqual(sheet["A21"].value, "below")

    def test_empty_work_items_min_rows(self) -> None:
        data = _workbook_bytes()
        payload = _payload()
        payload["content"]["workItems"] = []
        payload["proposal"]["amount"] = 0
        mapping = _mapping()
        mapping["tableBindings"][0]["minRows"] = 1
        with tempfile.TemporaryDirectory() as tmp:
            generated = render_custom_xlsx(
                template_bytes=data,
                payload=payload,
                mapping=mapping,
                output_dir=Path(tmp),
                output_file_name="out.xlsx",
            )
            workbook = load_workbook(generated.path)
            sheet = workbook["КП"]
            self.assertIsNone(sheet["B15"].value)

    def test_invalid_mapping_fails(self) -> None:
        data = _workbook_bytes()
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(DocumentGenerationError):
                render_custom_xlsx(
                    template_bytes=data,
                    payload=_payload(),
                    mapping={"schemaVersion": "1.0", "scalarBindings": [], "tableBindings": []},
                    output_dir=Path(tmp),
                    output_file_name="out.xlsx",
                )

    def test_rejects_non_xlsx(self) -> None:
        with self.assertRaises(DocumentGenerationError):
            inspect_xlsx(
                template_file_base64=base64.b64encode(b"not-a-zip").decode("ascii"),
                original_file_name="demo.xlsx",
            )

    def test_store_xlsx_template(self) -> None:
        data = _workbook_bytes()
        with tempfile.TemporaryDirectory() as tmp:
            storage = LocalDocumentStorage(Path(tmp) / "storage", "")
            result = store_xlsx_template(
                template_file_base64=base64.b64encode(data).decode("ascii"),
                original_file_name="demo.xlsx",
                storage=storage,
                expected_sha256=hashlib.sha256(data).hexdigest(),
            )
            self.assertEqual(result["status"], "success")
            self.assertEqual(result["sha256"], hashlib.sha256(data).hexdigest())
            self.assertTrue(result["storageKey"].startswith("xlsx-templates/"))
            self.assertTrue(storage.exists(storage_key=result["storageKey"]))
            stored = storage.get_bytes(storage_key=result["storageKey"])
            self.assertEqual(stored, data)
            self.assertNotIn("preview", result["workbook"]["sheets"][0])

    def test_load_by_storage_key_and_sha256(self) -> None:
        from mikoton_document_service.custom_xlsx import load_template_bytes_from_render_config

        data = _workbook_bytes()
        with tempfile.TemporaryDirectory() as tmp:
            storage = LocalDocumentStorage(Path(tmp) / "storage", "")
            stored = store_xlsx_template(
                template_file_base64=base64.b64encode(data).decode("ascii"),
                original_file_name="demo.xlsx",
                storage=storage,
            )
            loaded = load_template_bytes_from_render_config(
                {
                    "templateSource": "custom-xlsx",
                    "templateFile": {
                        "storageKey": stored["storageKey"],
                        "sha256": stored["sha256"],
                        "originalFileName": "demo.xlsx",
                    },
                },
                storage=storage,
            )
            self.assertEqual(loaded, data)

    def test_rejects_vertical_merge_on_template_row(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "КП"
        sheet["B15"] = "item"
        sheet.merge_cells("A14:A16")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "merged.xlsx"
            workbook.save(path)
            data = path.read_bytes()
            with self.assertRaises(DocumentGenerationError):
                render_custom_xlsx(
                    template_bytes=data,
                    payload=_payload(),
                    mapping=_mapping(),
                    output_dir=Path(tmp),
                    output_file_name="out.xlsx",
                )


if __name__ == "__main__":
    unittest.main()
